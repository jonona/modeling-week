import sys, os
os.system("pip3 install openpyxl") 
sys.path.insert(0, os.path.expanduser('~/.local/lib/python3.5/site-packages'))
from openpyxl import load_workbook
import numpy as np

## Function loads Data from the excel sheet Data.xlsx and writes its contents into numpy matrices/vectors
# The first dimension index is always the orientation number (-1), the rest depends on the variable dimensions:
# Tensors - (22,3,3)
# Eigenvectors - (22,3,3) where the  columns of the 3x3 matrix correspond to the eigenvectors
# Eigenvalues - (22,3)
# f_calc, f_eigval, g_calc, g_eigval, cn_calc, cn_eigval - (22,)
#
# To use in other script/notebook write:
# from loadData import *
# Tensors, Eigenvectors, Eigenvalues, f_calc, f_eigval, g_calc, g_eigval, cn_calc, cn_eigval = loadData(file_name, angle_switch)
# and set angle_switch=True for extended Data set and False for original table

def loadData(file_name="Data_extended.xlsx", angle_switch=False):
    wb = load_workbook(filename=file_name)
    sheet = wb.active

    n=22

    #initialize vectors
    Tensors = np.zeros((n,3,3))
    Eigenvectors = np.zeros((n,3,3))
    Eigenvalues = np.zeros((n,3))
    f_calc = np.zeros((n,))
    f_eigval = np.zeros((n,))
    g_calc = np.zeros((n,))
    g_eigval = np.zeros((n,))
    cn_calc = np.zeros((n,))
    cn_eigval = np.zeros((n,))

    if angle_switch:
        angle_calc = np.zeros((n,))
        angle_eigval = np.zeros((n,))
        #Read in vectors from excel sheet
        for j in range(1,n+1):
            f_calc[j-1] = sheet.cell(row=4*(j-1)+1,column=16).value
            f_eigval[j-1] = sheet.cell(row=4*(j-1)+1,column=18).value
            g_calc[j-1] = sheet.cell(row=4*(j-1)+1,column=20).value
            g_eigval[j-1] = sheet.cell(row=4*(j-1)+1,column=22).value
            cn_calc[j-1] = sheet.cell(row=4*(j-1)+1,column=24).value
            cn_eigval[j-1] = sheet.cell(row=4*(j-1)+1,column=26).value
            angle_calc[j-1] = sheet.cell(row=4*(j-1)+1,column=28).value
            angle_eigval[j-1] = sheet.cell(row=4*(j-1)+1,column=30).value
            for i in range(1,4):
                Eigenvalues[j-1,i-1] = sheet.cell(row=4*(j-1)+i,column=14).value
                for k in range(1,4):
                    Tensors[j-1, i-1,k-1] = sheet.cell(row=4*(j-1)+i,column=3+k).value
                    Eigenvectors[j-1,i-1,k-1] = sheet.cell(row=4*(j-1)+i,column=6+2*k).value

        return Tensors, Eigenvectors, Eigenvalues, f_calc, f_eigval, g_calc, g_eigval, cn_calc, cn_eigval, angle_calc, angle_eigval
    
    else:
        #Read in vectors from excel sheet
        for j in range(1,n+1):
            f_calc[j-1] = sheet.cell(row=4*(j-1)+1,column=16).value
            f_eigval[j-1] = sheet.cell(row=4*(j-1)+1,column=18).value
            g_calc[j-1] = sheet.cell(row=4*(j-1)+1,column=20).value
            g_eigval[j-1] = sheet.cell(row=4*(j-1)+1,column=22).value
            cn_calc[j-1] = sheet.cell(row=4*(j-1)+1,column=24).value
            cn_eigval[j-1] = sheet.cell(row=4*(j-1)+1,column=26).value
            for i in range(1,4):
                Eigenvalues[j-1,i-1] = sheet.cell(row=4*(j-1)+i,column=14).value
                for k in range(1,4):
                    Tensors[j-1, i-1,k-1] = sheet.cell(row=4*(j-1)+i,column=3+k).value
                    Eigenvectors[j-1,i-1,k-1] = sheet.cell(row=4*(j-1)+i,column=6+2*k).value

        return Tensors, Eigenvectors, Eigenvalues, f_calc, f_eigval, g_calc, g_eigval, cn_calc, cn_eigval